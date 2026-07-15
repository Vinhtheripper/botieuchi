import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook
from .database import connect

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DOCS_DIR = PROJECT_ROOT / "docs"
WORKBOOK = Path(os.getenv("SURVEY_WORKBOOK", DOCS_DIR / "BoCauHoi.xlsx"))
CRITERIA_WORKBOOK = Path(os.getenv("CRITERIA_WORKBOOK", DOCS_DIR / "BỘ TIÊU CHÍ.xlsx"))

def clean(value):
    if value is None: return None
    if isinstance(value, float) and value.is_integer(): return int(value)
    return value

def import_workbook(path: Path = WORKBOOK):
    wb = load_workbook(path, data_only=False)
    now = datetime.now(timezone.utc).isoformat()
    with connect() as con:
        # Workbook chính thức là source of truth: loại bỏ cấu hình từ bản cũ,
        # nhưng không đụng vào lịch sử respondent/answer đã thu thập.
        con.execute("DELETE FROM excel_sheets")
        con.execute("DELETE FROM variables")
        con.execute("DELETE FROM questions")
        for ws in wb.worksheets:
            matrix = [[clean(c.value) for c in row] for row in ws.iter_rows()]
            matrix = [[*row[:max((i for i, v in enumerate(row) if v is not None), default=-1)+1]] for row in matrix]
            while matrix and not any(v is not None for v in matrix[-1]): matrix.pop()
            con.execute("INSERT OR REPLACE INTO excel_sheets VALUES (?,?,?)", (ws.title, json.dumps(matrix, ensure_ascii=False), now))

        ws = wb["1_Biến nghiên cứu"]
        for r in range(4, 28):
            code = ws.cell(r, 1).value
            if not code or str(code).startswith("BIẾN"): continue
            vals = [clean(ws.cell(r, c).value) for c in range(1, 7)]
            con.execute("""INSERT INTO variables(code,name_vi,name_en,group_name,skip_rule,channel,active) VALUES(?,?,?,?,?,?,1)
                ON CONFLICT(code) DO UPDATE SET name_vi=excluded.name_vi,name_en=excluded.name_en,group_name=excluded.group_name,skip_rule=excluded.skip_rule,channel=excluded.channel""", vals)

        ws = wb["3_Bộ câu hỏi chính"]
        phase = "Trước mua"
        pos = 0
        for r in range(1, ws.max_row + 1):
            qid = ws.cell(r, 1).value
            if isinstance(qid, str) and qid.startswith("▶"):
                if "MỞ ĐẦU" in qid: phase = "Mở đầu"
                elif "ATTENTION" in qid: phase = "Kiểm tra tập trung"
                elif "TRƯỚC MUA" in qid: phase = "Trước mua"
                elif "ĐANG MUA" in qid: phase = "Đang mua"
                elif "CHUYỂN" in qid: phase = "Cá nhân hoá"
                elif "MUA LẶP" in qid: phase = "Mua lặp lại"
            if not isinstance(qid, str) or not re.fullmatch(r"(?:Q\d{2}|P\d{2}[ab]?|AC\d+)", qid): continue
            pos += 1
            text = ws.cell(r, 2).value or ""
            variables = [ws.cell(r, c).value for c in range(3, 6) if ws.cell(r, c).value]
            options = []
            rr = r + 1
            while rr <= ws.max_row and ws.cell(rr, 1).value in ("A", "B", "C", "D"):
                scores = {variables[i]: float(ws.cell(rr, 3+i).value) for i in range(len(variables)) if ws.cell(rr, 3+i).value is not None}
                options.append({"id": ws.cell(rr,1).value, "label": ws.cell(rr,2).value, "scores": scores})
                rr += 1
            if qid == "P00": kind = "text"
            elif qid.startswith("P"): kind = "piping"
            elif qid.startswith("AC"): kind = "attention"
            else: kind = "scenario"
            con.execute("""INSERT INTO questions(id,position,phase,kind,text,variables_json,options_json,note,active) VALUES(?,?,?,?,?,?,?,?,1)
                ON CONFLICT(id) DO UPDATE SET position=excluded.position,phase=excluded.phase,kind=excluded.kind,text=excluded.text,variables_json=excluded.variables_json,options_json=excluded.options_json,note=excluded.note""",
                (qid, pos, phase, kind, text, json.dumps(variables), json.dumps(options, ensure_ascii=False), ws.cell(r,6).value))

        def insert_after(after_id,qid,phase,kind,text,options):
            anchor=con.execute("SELECT position FROM questions WHERE id=?",(after_id,)).fetchone()
            if not anchor:return
            at=anchor["position"]+1
            con.execute("UPDATE questions SET position=position+1 WHERE position>=?",(at,))
            con.execute("INSERT INTO questions(id,position,phase,kind,text,variables_json,options_json,note,active) VALUES(?,?,?,?,?,'[]',?,'Cá nhân hóa trải nghiệm; không chấm điểm',1)",(qid,at,phase,kind,text,json.dumps(options,ensure_ascii=False)))
        continue_option=[{"id":"A","label":"Tiếp tục hành trình","scores":{}}]
        insert_after("Q12","T01","Chuyển chặng","transition","Mình đã bắt được một chút nhịp mua sắm của bạn.",continue_option)
        theme_options=[
            {"id":"A","label":"🌸 Hồng mận — ngọt nhưng có gu","theme":"rose","scores":{}},
            {"id":"B","label":"🌿 Mint cacao — dịu và tỉnh","theme":"mint","scores":{}},
            {"id":"C","label":"🌅 Cam hoàng hôn — ấm và nổi bật","theme":"sunset","scores":{}},
            {"id":"D","label":"🔮 Lavender đêm — mơ và cá tính","theme":"lavender","scores":{}},
        ]
        insert_after("P00","P01c","Mở đầu","theme","Trước khi bắt đầu, chọn một “mood” giao diện hợp gu bạn nhất?",theme_options)
        insert_after("P02","T02","Chuyển chặng","transition","Giờ thử đặt những lựa chọn vừa rồi vào món bạn thật sự đang quan tâm.",continue_option)
        pos=con.execute("SELECT MAX(position) max_position FROM questions").fetchone()["max_position"]

        ws = wb["4_Likert-mini & FIN"]
        for r in range(4, 10):
            qid = ws.cell(r,1).value
            if not qid: continue
            pos += 1
            variable = ws.cell(r,2).value
            text = ws.cell(r,3).value
            if qid == "F01":
                statement, option_text = text.split("  A.", 1)
                parts = re.split(r"\s+[BCD]\.\s+", option_text)
                options = [{"id": chr(65+i), "label": p.strip(), "scores": {variable: i+1}} for i,p in enumerate(parts)]
                text, kind = statement, "demographic"
            else:
                labels = [(ws.cell(rr,1).value, ws.cell(rr,2).value, ws.cell(rr,3).value) for rr in range(13,18)]
                options = [{"id": str(i), "label": f"{icon} {label}", "scores": {variable: i}} for i,icon,label in labels]
                kind = "likert"
            con.execute("""INSERT INTO questions(id,position,phase,kind,text,variables_json,options_json,note,active) VALUES(?,?,?,?,?,?,?,?,1)
                ON CONFLICT(id) DO UPDATE SET position=excluded.position,phase=excluded.phase,kind=excluded.kind,text=excluded.text,variables_json=excluded.variables_json,options_json=excluded.options_json,note=excluded.note""",
                (qid,pos,"Kết phiên",kind,text,json.dumps([variable]),json.dumps(options,ensure_ascii=False),ws.cell(r,4).value))
        demographics = [
            ("D01","Độ tuổi của bạn?",["Dưới 18","18–20","21–23","24–26","Trên 26"]),
            ("D02","Giới tính của bạn?",["Nữ","Nam","Phi nhị nguyên / Khác","Không muốn trả lời"]),
            ("D03","Nghề nghiệp hiện tại của bạn?",["Học sinh / Sinh viên","Nhân viên văn phòng","Kinh doanh / Tự do","Khác"]),
            ("D04","Bạn mua sắm online với tần suất nào?",["Ít hơn 1 lần/tháng","1–3 lần/tháng","1–2 lần/tuần","Từ 3 lần/tuần"]),
        ]
        for qid,text,labels in demographics:
            pos += 1
            options=[{"id":chr(65+i),"label":label,"scores":{}} for i,label in enumerate(labels)]
            con.execute("INSERT INTO questions(id,position,phase,kind,text,variables_json,options_json,note,active) VALUES(?,?,?,?,?,?,?,?,1)",(qid,pos,"Nhân khẩu học","demographic",text,"[]",json.dumps(options,ensure_ascii=False),"Theo spec sheet 5: tuổi/giới/nghề/tần suất mua"))
        con.execute("INSERT OR REPLACE INTO settings VALUES('last_import',?)", (now,))
        import_criteria(con, now)
    return {"sheets": wb.sheetnames, "imported_at": now}

def import_criteria(con, imported_at):
    if not CRITERIA_WORKBOOK.exists(): return
    wb=load_workbook(CRITERIA_WORKBOOK,data_only=False)
    for ws in wb.worksheets:
        matrix=[[clean(c.value) for c in row] for row in ws.iter_rows()]
        matrix=[row[:max((i for i,v in enumerate(row) if v is not None),default=-1)+1] for row in matrix]
        while matrix and not matrix[-1]: matrix.pop()
        name=f"TIÊU CHÍ · {ws.title}"
        con.execute("INSERT OR REPLACE INTO excel_sheets VALUES(?,?,?)",(name,json.dumps(matrix,ensure_ascii=False),imported_at))
    ws=wb["Tương quan"]; correlations={}; moderators=[]
    for r in range(3,ws.max_row+1):
        code=ws.cell(r,3).value
        if not code or not isinstance(code,str) or code == "Mã biến": continue
        role=ws.cell(r,2).value or ""
        if role=="Mod":
            moderators.append({"code":code,"relationship":ws.cell(r,6).value,"direction":ws.cell(r,7).value,"explanation":ws.cell(r,8).value})
        elif code!="PI":
            direction=str(ws.cell(r,6).value or "")
            correlations[code]={"sign":-1 if "Âm" in direction else 1,"direction":direction,"explanation":ws.cell(r,8).value}
    con.execute("INSERT OR REPLACE INTO settings VALUES('correlations',?)",(json.dumps(correlations,ensure_ascii=False),))
    con.execute("INSERT OR REPLACE INTO settings VALUES('moderator_relations',?)",(json.dumps(moderators,ensure_ascii=False),))
    # Mỗi tương quan có độ lớn mặc định 1; dấu lấy trực tiếp từ sheet Tương quan.
    con.execute("INSERT OR REPLACE INTO settings VALUES('heuristic_weights',?)",(json.dumps({k:v['sign'] for k,v in correlations.items()}),))
