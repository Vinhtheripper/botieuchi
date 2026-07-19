import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app import database
from app.database import init_db, row, rows
from app.excel_import import CRITERIA_WORKBOOK, WORKBOOK, import_workbook


class WorkbookIntegrityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        database.DB_PATH = Path(cls.tmp.name) / "workbook-integrity.db"
        init_db()
        import_workbook()

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def test_every_workbook_sheet_is_imported(self):
        expected = set(load_workbook(WORKBOOK, read_only=True).sheetnames)
        actual = {item["name"] for item in rows("SELECT name FROM excel_sheets")}
        self.assertTrue(expected.issubset(actual))

        criteria = {f"TIÊU CHÍ · {name}" for name in load_workbook(CRITERIA_WORKBOOK, read_only=True).sheetnames}
        self.assertTrue(criteria.issubset(actual))

    def test_question_variables_exist_and_scores_match_variables(self):
        variable_codes = {item["code"] for item in rows("SELECT code FROM variables")}
        self.assertGreaterEqual(len(variable_codes), 20)
        for question in rows("SELECT id,variables_json,options_json FROM questions"):
            variables = json.loads(question["variables_json"])
            self.assertFalse(set(variables) - variable_codes, question["id"])
            for option in json.loads(question["options_json"]):
                scores = option.get("scores", {})
                self.assertFalse(set(scores) - set(variables), f'{question["id"]}:{option.get("id")}')
                self.assertTrue(all(1 <= float(value) <= 5 for value in scores.values()), question["id"])

    def test_question_ids_are_unique_and_positions_are_contiguous(self):
        questions = rows("SELECT id,position FROM questions ORDER BY position")
        self.assertEqual(len(questions), len({item["id"] for item in questions}))
        self.assertEqual([item["position"] for item in questions], list(range(1, len(questions) + 1)))

    def test_heuristic_weight_signs_follow_criteria_sheet(self):
        correlations = json.loads(row("SELECT value FROM settings WHERE key='correlations'")["value"])
        weights = json.loads(row("SELECT value FROM settings WHERE key='heuristic_weights'")["value"])
        self.assertEqual(set(weights), set(correlations))
        for code, relation in correlations.items():
            self.assertEqual(weights[code], relation["sign"], code)


if __name__ == "__main__":
    unittest.main()
